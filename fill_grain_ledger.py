import os
import re
import shutil
import tempfile
import openpyxl
from openpyxl import load_workbook
from copy import copy

# ── 配置 ─────────────────────────────────────────────────────────
TEMPLATE = 'C:/Users/neo/Desktop/米面粮油干货脚本/粮油干调台账模板.xls'
BILL = 'C:/Users/neo/Desktop/米面粮油干货脚本/山禾（松山湖）5月份对账单.xls'
OUTPUT_DIR = 'C:/Users/neo/Desktop/米面粮油干货脚本'


# ── 辅助：.xls → .xlsx 转换 ──────────────────────────────────────
def convert_xls_to_xlsx(src_path):
    """把 .xls 转为 .xlsx，优先 win32com（格式最完整），回退 xlrd+openpyxl。"""
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()

    try:
        import win32com.client
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(src_path))
        wb.SaveAs(tmp_path, FileFormat=51)   # 51 = xlOpenXMLWorkbook
        wb.Close(SaveChanges=False)
        excel.Quit()
        return tmp_path
    except Exception:
        # 回退：xlrd 读取 → openpyxl 重建
        import xlrd
        rb = xlrd.open_workbook(src_path)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 去掉默认 sheet
        for sheet_name in rb.sheet_names():
            ws_src = rb.sheet_by_name(sheet_name)
            ws_dst = wb.create_sheet(title=sheet_name)
            for r in range(ws_src.nrows):
                for c in range(ws_src.ncols):
                    ws_dst.cell(row=r+1, column=c+1, value=ws_src.cell_value(r, c))
        wb.save(tmp_path)
        return tmp_path


# ── 辅助：复制一行的格式 ──────────────────────────────────────────
def copy_row_format(ws, src_row, dst_row, max_col=7):
    """复制一行的格式（行高 + 单元格样式）"""
    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    if src_dim.height is not None:
        dst_dim.height = src_dim.height

    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = copy(src_cell.number_format)
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)


# ── 解析对账单 ────────────────────────────────────────────────────
def parse_bill(bill_path):
    """解析对账单，返回 {date_str: [数据字典列表]}"""
    import xlrd
    rb = xlrd.open_workbook(bill_path)
    ws = rb.sheet_by_index(0)

    data_by_date = {}
    current_date = None

    for r in range(ws.nrows):
        cell_a = str(ws.cell_value(r, 0)).strip()

        # 日期行（格式：YYYY-MM-DD）
        if re.match(r'^\d{4}-\d{2}-\d{2}$', cell_a):
            current_date = cell_a
            data_by_date[current_date] = []
            continue

        # 没有日期上下文则跳过
        if current_date is None:
            continue

        # 跳过汇总行（金额 / 小计 / 总计）
        if cell_a in ('金额', '总计') or '小计' in cell_a:
            continue

        # 跳过空行
        if cell_a == '':
            continue

        # 数据行：A列必须是序号（正整数）
        try:
            seq = int(float(cell_a))
            if seq <= 0:
                continue

            name = str(ws.cell_value(r, 1)).strip()
            if not name or name in ('品名', 'NaN', 'nan'):
                continue

            unit = str(ws.cell_value(r, 2)).strip() if ws.cell_value(r, 2) else ''
            qty = ws.cell_value(r, 3) if ws.cell_value(r, 3) else 0
            actual_price = ws.cell_value(r, 5) if ws.ncols > 5 and ws.cell_value(r, 5) else 0
            amount = ws.cell_value(r, 6) if ws.ncols > 6 and ws.cell_value(r, 6) else 0

            data_by_date[current_date].append({
                'seq': seq,
                'name': name,
                'unit': unit,
                'qty': qty,
                'price': actual_price,   # 下浮后的实际采购单价
                'amount': amount
            })
        except (ValueError, TypeError):
            continue

    rb.release_resources()
    return data_by_date


# ── 主程序 ────────────────────────────────────────────────────────
# 1. 模板转 xlsx
print('[INFO] 正在转换模板...')
tmp_template = convert_xls_to_xlsx(TEMPLATE)

# 2. 解析对账单
print('[INFO] 正在解析对账单...')
data_by_date = parse_bill(BILL)
print(f'共解析到 {len(data_by_date)} 个日期: {sorted(data_by_date.keys())}')
for d, items in sorted(data_by_date.items()):
    print(f'  {d}: {len(items)} 条')

# 3. 创建副本（基于模板）
bill_name = os.path.splitext(os.path.basename(BILL))[0]
dst_path = os.path.join(OUTPUT_DIR, bill_name + '输出结果.xlsx')
shutil.copy2(tmp_template, dst_path)
os.remove(tmp_template)
print(f'[INFO] 已复制模板 → {dst_path}')

# 4. 用 openpyxl 打开副本
wb = load_workbook(dst_path)
mother_sheet = wb.worksheets[0]
mother_name = mother_sheet.title

# 5. 为每个日期创建 sheet（基于母版复制）
for date_str, items in sorted(data_by_date.items()):
    new_sheet = wb.copy_worksheet(mother_sheet)
    new_sheet.title = date_str
    N = len(items)

    # 5a. 更新时间行（行2）
    new_sheet.cell(row=2, column=6).value = f'时间：{date_str.replace("-", ".")}'

    # 原模板：数据区行4~行23（20行），合计行24
    # 目标：数据区行4~行(3+N)（N行），合计行(4+N）
    if N > 20:
        # 在合计行前插入 N-20 行
        insert_pos = 24
        new_sheet.insert_rows(insert_pos, N - 20)
        for r in range(insert_pos, insert_pos + N - 20):
            copy_row_format(new_sheet, 23, r)
            # 填入序号
            new_sheet.cell(row=r, column=1).value = r - 3
    elif N < 20:
        # 删除多余的数据行：从行(4+N)到行23，共 20-N 行
        new_sheet.delete_rows(4 + N, 20 - N)
        # 修复合计行格式（delete_rows后继承了被删除行的格式，需还原）
        total_row = 4 + N
        for col in range(1, 8):
            src_cell = mother_sheet.cell(row=24, column=col)
            dst_cell = new_sheet.cell(row=total_row, column=col)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.border = copy(src_cell.border)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.number_format = copy(src_cell.number_format)
                dst_cell.protection = copy(src_cell.protection)
                dst_cell.alignment = copy(src_cell.alignment)
        if mother_sheet.row_dimensions[24].height is not None:
            new_sheet.row_dimensions[total_row].height = mother_sheet.row_dimensions[24].height

    total_row = 4 + N

    # 5b. 填写N行数据
    for i, item in enumerate(items):
        row_num = 4 + i
        new_sheet.cell(row=row_num, column=1).value = i + 1   # A 序号
        new_sheet.cell(row=row_num, column=2).value = item['name']   # B 品名
        new_sheet.cell(row=row_num, column=3).value = item['unit']   # C 单位
        new_sheet.cell(row=row_num, column=4).value = item['qty']    # D 数量
        new_sheet.cell(row=row_num, column=5).value = item['price']  # E 单价
        new_sheet.cell(row=row_num, column=6).value = item['amount'] # F 金额
        new_sheet.cell(row=row_num, column=7).value = None           # G 清空

    # 5c. 更新合计行
    new_sheet.cell(row=total_row, column=2).value = '小计金额'
    for c in [1, 3, 4, 5]:
        new_sheet.cell(row=total_row, column=c).value = None
    total_amount = sum(item['amount'] for item in items)
    new_sheet.cell(row=total_row, column=6).value = total_amount
    new_sheet.cell(row=total_row, column=7).value = None

    print(f'  Sheet {date_str}: {N} 条，小计 {total_amount:.2f}')

# 6. 删除母版 sheet
if mother_name in wb.sheetnames:
    del wb[mother_name]

# 7. 删除多余的模板 sheet（保留日期格式的 sheet）
for sheet_name in list(wb.sheetnames):
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', sheet_name):
        del wb[sheet_name]

# 8. 保存
wb.save(dst_path)
print(f'\n完成！')
print(f'  共生成 {len(data_by_date)} 个日期 sheet')
print(f'  输出文件: {dst_path}')
