import os
import re
import shutil
import tempfile
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl import load_workbook
from copy import copy

# ── 辅助：.xls → .xlsx 转换 ──────────────────────────────────────
def convert_xls_to_xlsx(src_path):
    """把 .xls 转为 .xlsx，优先 win32com（格式最完整），回退 xlrd+openpyxl。"""
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()

    try:
        import win32com.client
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(src_path))
        wb.SaveAs(tmp_path, FileFormat=51)
        wb.Close(SaveChanges=False)
        excel.Quit()
        return tmp_path
    except Exception:
        import xlrd
        rb = xlrd.open_workbook(src_path)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name in rb.sheet_names():
            ws_src = rb.sheet_by_name(sheet_name)
            ws_dst = wb.create_sheet(title=sheet_name)
            for r in range(ws_src.nrows):
                for c in range(ws_src.ncols):
                    ws_dst.cell(row=r+1, column=c+1, value=ws_src.cell_value(r, c))
        wb.save(tmp_path)
        return tmp_path


# ── 辅助：复制一行的格式 ──────────────────────────────────────────
def copy_row_format(ws, src_row, dst_row, max_col=7):
    """复制一行的格式（行高 + 单元格样式）"""
    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    if src_dim.height is not None:
        dst_dim.height = src_dim.height

    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = copy(src_cell.number_format)
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)


# ── 解析对账单 ────────────────────────────────────────────────────
def parse_bill(bill_path):
    """解析对账单，返回 {date_str: [数据字典列表]}"""
    import xlrd
    rb = xlrd.open_workbook(bill_path)
    ws = rb.sheet_by_index(0)

    data_by_date = {}
    current_date = None

    for r in range(ws.nrows):
        cell_a = str(ws.cell_value(r, 0)).strip()

        if re.match(r'^\d{4}-\d{2}-\d{2}$', cell_a):
            current_date = cell_a
            data_by_date[current_date] = []
            continue

        if current_date is None:
            continue

        if cell_a in ('金额', '总计') or '小计' in cell_a:
            continue

        if cell_a == '':
            continue

        try:
            seq = int(float(cell_a))
            if seq <= 0:
                continue

            name = str(ws.cell_value(r, 1)).strip()
            if not name or name in ('品名', 'NaN', 'nan'):
                continue

            unit = str(ws.cell_value(r, 2)).strip() if ws.cell_value(r, 2) else ''
            qty = ws.cell_value(r, 3) if ws.cell_value(r, 3) else 0
            actual_price = ws.cell_value(r, 5) if ws.ncols > 5 and ws.cell_value(r, 5) else 0
            amount = ws.cell_value(r, 6) if ws.ncols > 6 and ws.cell_value(r, 6) else 0

            data_by_date[current_date].append({
                'seq': seq,
                'name': name,
                'unit': unit,
                'qty': qty,
                'price': actual_price,
                'amount': amount
            })
        except (ValueError, TypeError):
            continue

    rb.release_resources()
    return data_by_date


# ── 核心处理 ──────────────────────────────────────────────────────
def process(template_path, bill_path, output_path, log_callback=None):
    """主处理逻辑"""
    def log(msg):
        if log_callback:
            log_callback(msg)

    # 1. 模板转 xlsx
    log('[INFO] 正在转换模板...')
    tmp_template = convert_xls_to_xlsx(template_path)

    # 2. 解析对账单
    log('[INFO] 正在解析对账单...')
    data_by_date = parse_bill(bill_path)
    log(f'共解析到 {len(data_by_date)} 个日期: {sorted(data_by_date.keys())}')
    for d, items in sorted(data_by_date.items()):
        log(f'  {d}: {len(items)} 条')

    # 3. 创建副本
    shutil.copy2(tmp_template, output_path)
    os.remove(tmp_template)
    log(f'[INFO] 已复制模板 → {output_path}')

    # 4. 用 openpyxl 打开副本
    wb = load_workbook(output_path)
    mother_sheet = wb.worksheets[0]
    mother_name = mother_sheet.title

    # 5. 为每个日期创建 sheet
    for date_str, items in sorted(data_by_date.items()):
        new_sheet = wb.copy_worksheet(mother_sheet)
        new_sheet.title = date_str
        N = len(items)

        new_sheet.cell(row=2, column=6).value = f'时间：{date_str.replace("-", ".")}'

        if N > 20:
            insert_pos = 24
            new_sheet.insert_rows(insert_pos, N - 20)
            for r in range(insert_pos, insert_pos + N - 20):
                copy_row_format(new_sheet, 23, r)
                new_sheet.cell(row=r, column=1).value = r - 3
        elif N < 20:
            new_sheet.delete_rows(4 + N, 20 - N)
            total_row = 4 + N
            for col in range(1, 8):
                src_cell = mother_sheet.cell(row=24, column=col)
                dst_cell = new_sheet.cell(row=total_row, column=col)
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.number_format = copy(src_cell.number_format)
                    dst_cell.protection = copy(src_cell.protection)
                    dst_cell.alignment = copy(src_cell.alignment)
            if mother_sheet.row_dimensions[24].height is not None:
                new_sheet.row_dimensions[total_row].height = mother_sheet.row_dimensions[24].height

        total_row = 4 + N

        for i, item in enumerate(items):
            row_num = 4 + i
            new_sheet.cell(row=row_num, column=1).value = i + 1
            new_sheet.cell(row=row_num, column=2).value = item['name']
            new_sheet.cell(row=row_num, column=3).value = item['unit']
            new_sheet.cell(row=row_num, column=4).value = item['qty']
            new_sheet.cell(row=row_num, column=5).value = item['price']
            new_sheet.cell(row=row_num, column=6).value = item['amount']
            new_sheet.cell(row=row_num, column=7).value = None

        new_sheet.cell(row=total_row, column=2).value = '小计金额'
        for c in [1, 3, 4, 5]:
            new_sheet.cell(row=total_row, column=c).value = None
        total_amount = sum(item['amount'] for item in items)
        new_sheet.cell(row=total_row, column=6).value = total_amount
        new_sheet.cell(row=total_row, column=7).value = None

        log(f'  Sheet {date_str}: {N} 条，小计 {total_amount:.2f}')

    # 6. 删除母版
    if mother_name in wb.sheetnames:
        del wb[mother_name]

    # 7. 删除多余的模板 sheet
    for sheet_name in list(wb.sheetnames):
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', sheet_name):
            del wb[sheet_name]

    # 8. 保存
    wb.save(output_path)
    log(f'\n完成！')
    log(f'  共生成 {len(data_by_date)} 个日期 sheet')
    log(f'  输出文件: {output_path}')
    return True


# ── GUI ───────────────────────────────────────────────────────────
class GrainLedgerGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('粮油干货台账自动填充')
        self.geometry('640x520')
        self.resizable(False, False)

        # 居中窗口
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        y = (self.winfo_screenheight() // 2) - (h // 2)
        self.geometry(f'+{x}+{y}')

        self._build_ui()

    def _build_ui(self):
        pad = dict(padx=12, pady=8)
        file_pad = dict(padx=12, pady=4)

        # ── 标题 ──
        ttk.Label(self, text='粮油干货台账自动填充', font=('微软雅黑', 16, 'bold')).pack(pady=(16, 8))

        # ── 模板文件 ──
        f1 = ttk.Frame(self)
        f1.pack(fill='x', **file_pad)
        ttk.Label(f1, text='台账模板：', width=10, anchor='e').pack(side='left')
        self.tpl_var = tk.StringVar(value='')
        ttk.Entry(f1, textvariable=self.tpl_var, width=50).pack(side='left', padx=(4, 4))
        ttk.Button(f1, text='选择文件', command=self._pick_tpl).pack(side='left')

        # ── 对账单文件 ──
        f2 = ttk.Frame(self)
        f2.pack(fill='x', **file_pad)
        ttk.Label(f2, text='对账单：', width=10, anchor='e').pack(side='left')
        self.bill_var = tk.StringVar()
        ttk.Entry(f2, textvariable=self.bill_var, width=50).pack(side='left', padx=(4, 4))
        ttk.Button(f2, text='选择文件', command=self._pick_bill).pack(side='left')

        # ── 输出路径 ──
        f3 = ttk.Frame(self)
        f3.pack(fill='x', **file_pad)
        ttk.Label(f3, text='输出路径：', width=10, anchor='e').pack(side='left')
        self.out_var = tk.StringVar()
        ttk.Entry(f3, textvariable=self.out_var, width=50).pack(side='left', padx=(4, 4))
        ttk.Button(f3, text='选择路径', command=self._pick_out).pack(side='left')

        # ── 运行按钮 ──
        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=12)
        self.run_btn = ttk.Button(btn_frame, text='开始处理', command=self._run, width=20)
        self.run_btn.pack()

        # ── 日志区域 ──
        ttk.Label(self, text='处理日志：').pack(anchor='w', **pad)
        log_frame = ttk.Frame(self)
        log_frame.pack(fill='both', expand=True, padx=12, pady=(0, 12))

        scrollbar = ttk.Scrollbar(log_frame)
        scrollbar.pack(side='right', fill='y')

        self.log_text = tk.Text(log_frame, height=14, wrap='word', yscrollcommand=scrollbar.set,
                                font=('Consolas', 10), state='disabled')
        self.log_text.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.log_text.yview)

    def _pick_tpl(self):
        path = filedialog.askopenfilename(
            title='选择台账模板',
            filetypes=[('Excel 文件', '*.xls *.xlsx')]
        )
        if path:
            self.tpl_var.set(path)
            self._auto_derive_out()

    def _pick_bill(self):
        path = filedialog.askopenfilename(
            title='选择对账单',
            filetypes=[('Excel 文件', '*.xls *.xlsx')]
        )
        if path:
            self.bill_var.set(path)
            self._auto_derive_out()

    def _pick_out(self):
        initial = self.out_var.get()
        if not initial:
            initial = 'C:/'
        path = filedialog.asksaveasfilename(
            title='保存输出文件',
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')],
            initialfile=os.path.basename(initial) if initial else '输出结果.xlsx',
            initialdir=os.path.dirname(initial) if initial else 'C:/'
        )
        if path:
            self.out_var.set(path)

    def _auto_derive_out(self):
        """根据对账单路径自动推导输出路径"""
        bill = self.bill_var.get()
        if not bill:
            return
        base, _ = os.path.splitext(bill)
        out = base + '输出结果.xlsx'
        if not self.out_var.get():
            self.out_var.set(out)

    def _log(self, msg):
        self.log_text.config(state='normal')
        self.log_text.insert('end', msg + '\n')
        self.log_text.see('end')
        self.log_text.config(state='disabled')
        self.update_idletasks()

    def _run(self):
        template = self.tpl_var.get().strip()
        bill = self.bill_var.get().strip()
        output = self.out_var.get().strip()

        if not template or not os.path.isfile(template):
            messagebox.showerror('错误', '请选择一个有效的台账模板文件')
            return
        if not bill or not os.path.isfile(bill):
            messagebox.showerror('错误', '请选择一个有效的对账单文件')
            return
        if not output:
            messagebox.showerror('错误', '请指定输出路径')
            return

        self.run_btn.config(state='disabled')
        self.log_text.config(state='normal')
        self.log_text.delete('1.0', 'end')
        self.log_text.config(state='disabled')

        try:
            process(template, bill, output, log_callback=self._log)
            messagebox.showinfo('完成', f'处理成功！\n输出文件: {output}')
        except Exception as e:
            import traceback
            self._log(traceback.format_exc())
            messagebox.showerror('错误', f'处理失败: {str(e)}')
        finally:
            self.run_btn.config(state='normal')


if __name__ == '__main__':
    app = GrainLedgerGUI()
    app.mainloop()
